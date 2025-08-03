from sqlalchemy.orm import Session
from sqlalchemy import and_, or_
from typing import List, Optional, Dict, Any
from datetime import date, datetime, timedelta
from decimal import Decimal
import openpyxl
from openpyxl import load_workbook
import io
import csv
import pandas as pd
import xlrd
import logging
import hashlib
from ..utils.transaction_utils import generate_transaction_hash, check_transaction_exists

from ..models.transactions import Transaction, TransactionStatus
from ..models.accounts import Account
from ..models.file_imports import FileImportProfile, FileColumnMapping
from ..schemas.transactions import TransactionCreateRequest, TransactionUpdateRequest

# Configurar logger
logger = logging.getLogger(__name__)

def get_default_transaction_status_id(db: Session):
    """Obtiene el ID de estado de transacción por defecto, creándolo si no existe"""
    
    # Buscar un estado existente
    status = db.query(TransactionStatus).first()
    
    if status:
        status_id = status.id
        logger.debug(f"Estado de transacción existente encontrado: {status_id} - {status.name}")
        return status_id
    
    # Si no existe ningún estado, crear uno por defecto
    logger.warning("No se encontraron estados de transacción, creando estado por defecto")
    default_status = TransactionStatus(
        id=1,
        name="Completada",
        description="Transacción completada exitosamente"
    )
    
    try:
        db.add(default_status)
        db.commit()
        db.refresh(default_status)
        status_id = default_status.id
        logger.info(f"Estado de transacción por defecto creado: {status_id} - {default_status.name}")
        return status_id
    except Exception as e:
        logger.error(f"Error creando estado por defecto: {str(e)}")
        db.rollback()
        # Si falla la creación, intentar obtener uno existente nuevamente
        existing_status = db.query(TransactionStatus).first()
        if existing_status:
            return existing_status.id
        else:
            raise ValueError("No se pudo crear ni obtener un estado de transacción válido")

def create_transaction(
    db: Session, 
    user_id: int, 
    transaction_data: TransactionCreateRequest,
    import_source: Optional[str] = None,
    skip_duplicate_check: bool = False
) -> Transaction:
    """Crea una nueva transacción con validación de duplicados"""
    
    logger.info(f"Iniciando creación de transacción para usuario {user_id}")
    logger.debug(f"Datos de transacción: amount={transaction_data.amount}, account_id={transaction_data.account_id}, description='{transaction_data.description}'")
    
    # Verificar que la cuenta pertenece al usuario
    account = db.query(Account).filter(
        Account.id == transaction_data.account_id,
        Account.user_id == user_id,
        Account.active == True
    ).first()
    
    if not account:
        logger.error(f"Cuenta {transaction_data.account_id} no encontrada o no autorizada para usuario {user_id}")
        raise ValueError("Cuenta no encontrada o no autorizada")
    
    # NUEVA VALIDACIÓN DE DUPLICADOS
    if not skip_duplicate_check:
        logger.debug("Verificando duplicados...")
        
        # Generar hash del contenido
        content_hash = generate_transaction_hash(
            user_id=user_id,
            account_id=transaction_data.account_id,
            amount=Decimal(str(transaction_data.amount)),
            description=transaction_data.description or '',
            transaction_date=transaction_data.transaction_date,
            external_id=transaction_data.external_id
        )
        
        # Verificar si ya existe
        exists, reason = check_transaction_exists(
            db=db,
            user_id=user_id,
            account_id=transaction_data.account_id,
            amount=Decimal(str(transaction_data.amount)),
            description=transaction_data.description or '',
            transaction_date=transaction_data.transaction_date,
            external_id=transaction_data.external_id,
            content_hash=content_hash
        )
        
        if exists:
            logger.warning(f"Transacción duplicada detectada: {reason}")
            raise ValueError(f"Transacción duplicada: {reason}")
        
        logger.debug(f"No se encontraron duplicados, content_hash: {content_hash[:16]}...")
    else:
        content_hash = None
        logger.debug("Verificación de duplicados saltada")
    
    # Asegurar que el monto sea un Decimal con precisión adecuada
    amount_decimal = Decimal(str(transaction_data.amount)).quantize(Decimal('0.01'))
    
    # Obtener un status_id válido
    status_id = transaction_data.status_id
    if not status_id:
        status_id = get_default_transaction_status_id(db)
        logger.debug(f"Usando status_id por defecto: {status_id}")
    
    # Crear la transacción con los nuevos campos
    db_transaction = Transaction(
        user_id=user_id,
        account_id=transaction_data.account_id,
        amount=amount_decimal,
        description=transaction_data.description,
        notes=transaction_data.notes,
        transaction_date=transaction_data.transaction_date,
        transfer_account_id=transaction_data.transfer_account_id,
        subcategory_id=transaction_data.subcategory_id,
        envelope_id=transaction_data.envelope_id,
        status_id=status_id,
        is_recurring=transaction_data.is_recurring,
        is_planned=transaction_data.is_planned,
        kakebo_emotion=transaction_data.kakebo_emotion,
        external_id=transaction_data.external_id,
        # NUEVOS CAMPOS
        content_hash=content_hash,
        import_source=import_source
    )
    
    db.add(db_transaction)
    logger.debug(f"Transacción agregada a la sesión de BD")
    
    # Actualizar balance de la cuenta
    old_balance = account.current_balance
    account.current_balance += float(transaction_data.amount)  # type: ignore
    logger.debug(f"Balance de cuenta {account.id} actualizado: {old_balance} -> {account.current_balance}")
    
    # Si es transferencia, actualizar cuenta destino
    if transaction_data.transfer_account_id and transaction_data.amount < 0:
        transfer_account = db.query(Account).filter(Account.id == transaction_data.transfer_account_id).first()
        if transfer_account:
            old_transfer_balance = transfer_account.current_balance
            transfer_account.current_balance += abs(float(transaction_data.amount))  # type: ignore
            logger.debug(f"Balance de cuenta de transferencia {transfer_account.id} actualizado: {old_transfer_balance} -> {transfer_account.current_balance}")
    
    try:
        db.commit()
        logger.info(f"Transacción creada exitosamente con ID {db_transaction.id}")
    except Exception as e:
        logger.error(f"Error al hacer commit de la transacción: {str(e)}")
        db.rollback()
        raise
    
    db.refresh(db_transaction)
    
    return db_transaction

def get_user_transactions(
    db: Session, 
    user_id: int,
    account_id: Optional[int] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    subcategory_id: Optional[int] = None,
    skip: int = 0,
    limit: int = 50
) -> List[Transaction]:
    """Obtiene las transacciones del usuario con filtros"""
    
    logger.info(f"Obteniendo transacciones para usuario {user_id}")
    logger.debug(f"Filtros: account_id={account_id}, start_date={start_date}, end_date={end_date}, subcategory_id={subcategory_id}, skip={skip}, limit={limit}")
    
    query = db.query(Transaction).filter(Transaction.user_id == user_id)
    
    if account_id:
        query = query.filter(Transaction.account_id == account_id)
        logger.debug(f"Filtro aplicado: account_id={account_id}")
    
    if start_date:
        query = query.filter(Transaction.transaction_date >= start_date)
        logger.debug(f"Filtro aplicado: start_date={start_date}")
    
    if end_date:
        query = query.filter(Transaction.transaction_date <= end_date)
        logger.debug(f"Filtro aplicado: end_date={end_date}")
    
    if subcategory_id:
        query = query.filter(Transaction.subcategory_id == subcategory_id)
        logger.debug(f"Filtro aplicado: subcategory_id={subcategory_id}")
    
    query = query.order_by(Transaction.transaction_date.desc(), Transaction.id.desc())
    query = query.offset(skip).limit(limit)
    
    try:
        transactions = query.all()
        logger.info(f"Se encontraron {len(transactions)} transacciones para usuario {user_id}")
        return transactions
    except Exception as e:
        logger.error(f"Error al obtener transacciones para usuario {user_id}: {str(e)}")
        raise

def get_user_transaction(db: Session, user_id: int, transaction_id: int) -> Optional[Transaction]:
    """Obtiene una transacción específica del usuario"""
    logger.debug(f"Buscando transacción {transaction_id} para usuario {user_id}")
    
    transaction = db.query(Transaction).filter(
        Transaction.id == transaction_id,
        Transaction.user_id == user_id
    ).first()
    
    if transaction:
        logger.debug(f"Transacción {transaction_id} encontrada")
    else:
        logger.warning(f"Transacción {transaction_id} no encontrada para usuario {user_id}")
    
    return transaction

def update_transaction(
    db: Session, 
    user_id: int, 
    transaction_id: int, 
    transaction_data: TransactionUpdateRequest
) -> Optional[Transaction]:
    """Actualiza una transacción existente"""
    
    logger.info(f"Iniciando actualización de transacción {transaction_id} para usuario {user_id}")
    logger.debug(f"Datos de actualización: {transaction_data.dict(exclude_unset=True)}")
    
    # Obtener la transacción
    db_transaction = db.query(Transaction).filter(
        Transaction.id == transaction_id,
        Transaction.user_id == user_id
    ).first()
    
    if not db_transaction:
        logger.warning(f"Transacción {transaction_id} no encontrada para usuario {user_id}")
        return None
    
    logger.debug(f"Transacción encontrada: amount={db_transaction.amount}, account_id={db_transaction.account_id}")
    
    # Guardar monto anterior para ajustar balances
    old_amount = db_transaction.amount
    old_account_id = db_transaction.account_id
    
    # Actualizar campos que no son None
    update_data = transaction_data.dict(exclude_unset=True)
    
    for field, value in update_data.items():
        if hasattr(db_transaction, field):
            # Convertir amount a Decimal con precisión adecuada
            if field == 'amount' and value is not None:
                value = Decimal(str(value)).quantize(Decimal('0.01'))
            logger.debug(f"Actualizando campo {field}: {getattr(db_transaction, field)} -> {value}")
            setattr(db_transaction, field, value)
    
    db_transaction.updated_at = datetime.utcnow()
    
    # Ajustar balances si cambió el monto o la cuenta
    if 'amount' in update_data or 'account_id' in update_data:
        logger.debug("Ajustando balances de cuentas por cambio de monto o cuenta")
        
        # Revertir el monto anterior de la cuenta anterior
        old_account = db.query(Account).filter(Account.id == old_account_id).first()
        if old_account:
            old_balance = old_account.current_balance
            old_account.current_balance -= float(old_amount)  # type: ignore
            logger.debug(f"Balance cuenta anterior {old_account_id}: {old_balance} -> {old_account.current_balance}")
        
        # Aplicar nuevo monto a la cuenta (nueva o misma)
        new_account_id = update_data.get('account_id', old_account_id)
        new_amount = float(update_data.get('amount', old_amount))
        
        new_account = db.query(Account).filter(Account.id == new_account_id).first()
        if new_account:
            new_balance = new_account.current_balance
            new_account.current_balance += new_amount  # type: ignore
            logger.debug(f"Balance cuenta nueva {new_account_id}: {new_balance} -> {new_account.current_balance}")
    
    try:
        db.commit()
        logger.info(f"Transacción {transaction_id} actualizada exitosamente")
    except Exception as e:
        logger.error(f"Error al actualizar transacción {transaction_id}: {str(e)}")
        db.rollback()
        raise
    
    db.refresh(db_transaction)
    
    return db_transaction

def delete_transaction(db: Session, user_id: int, transaction_id: int) -> bool:
    """Elimina una transacción"""
    
    logger.info(f"Iniciando eliminación de transacción {transaction_id} para usuario {user_id}")
    
    db_transaction = db.query(Transaction).filter(
        Transaction.id == transaction_id,
        Transaction.user_id == user_id
    ).first()
    
    if not db_transaction:
        logger.warning(f"Transacción {transaction_id} no encontrada para usuario {user_id}")
        return False
    
    logger.debug(f"Transacción encontrada: amount={db_transaction.amount}, account_id={db_transaction.account_id}")
    
    # Revertir el balance de la cuenta
    account = db.query(Account).filter(Account.id == db_transaction.account_id).first()
    if account:
        old_balance = account.current_balance
        account.current_balance -= float(db_transaction.amount)  # type: ignore
        logger.debug(f"Balance cuenta {account.id} revertido: {old_balance} -> {account.current_balance}")
    
    # Si era transferencia, revertir cuenta destino
    if db_transaction.transfer_account_id and db_transaction.amount < 0:  # type: ignore
        transfer_account = db.query(Account).filter(Account.id == db_transaction.transfer_account_id).first()
        if transfer_account:
            old_transfer_balance = transfer_account.current_balance
            transfer_account.current_balance -= abs(float(db_transaction.amount))  # type: ignore
            logger.debug(f"Balance cuenta transferencia {transfer_account.id} revertido: {old_transfer_balance} -> {transfer_account.current_balance}")
    
    try:
        db.delete(db_transaction)
        db.commit()
        logger.info(f"Transacción {transaction_id} eliminada exitosamente")
        return True
    except Exception as e:
        logger.error(f"Error al eliminar transacción {transaction_id}: {str(e)}")
        db.rollback()
        raise

def parse_excel_date(value):
    """Convierte diferentes formatos de fecha de Excel a date"""
    if value is None:
        return None
    
    logger.debug(f"Parseando fecha: {value} (tipo: {type(value)})")
    
    # Si ya es un objeto datetime
    if isinstance(value, datetime):
        result = value.date()
        logger.debug(f"Fecha convertida desde datetime: {result}")
        return result
    
    # Si ya es un objeto date
    if isinstance(value, date):
        logger.debug(f"Fecha ya es date: {value}")
        return value
    
    # Si es string, intentar parsear diferentes formatos
    if isinstance(value, str):
        date_formats = [
            '%Y-%m-%d',      # 2024-01-15
            '%d/%m/%Y',      # 15/01/2024
            '%d-%m-%Y',      # 15-01-2024
            '%m/%d/%Y',      # 01/15/2024
            '%d/%m/%y',      # 15/01/24
            '%Y%m%d',        # 20240115
        ]
        
        for fmt in date_formats:
            try:
                result = datetime.strptime(value.strip(), fmt).date()
                logger.debug(f"Fecha parseada con formato {fmt}: {result}")
                return result
            except ValueError:
                continue
        
        logger.error(f"Formato de fecha no válido: {value}")
        raise ValueError(f"Formato de fecha no válido: {value}")
    
    # Si es un número (serial date de Excel)
    if isinstance(value, (int, float)):
        try:
            # Excel cuenta días desde 1900-01-01 (con ajuste por bug de año bisiesto)
            excel_epoch = datetime(1899, 12, 30)
            result = (excel_epoch + timedelta(days=value)).date()
            logger.debug(f"Fecha convertida desde serial Excel {value}: {result}")
            return result
        except Exception as e:
            logger.error(f"Error convirtiendo fecha numérica {value}: {str(e)}")
            raise ValueError(f"Fecha numérica no válida: {value}")
    
    logger.error(f"Tipo de fecha no soportado: {type(value)}")
    raise ValueError(f"Tipo de fecha no soportado: {type(value)}")

def parse_excel_amount(value):
    """Convierte valores de Excel a float"""
    if value is None or value == '':
        return 0.0
    
    logger.debug(f"Parseando monto: {value} (tipo: {type(value)})")
    
    if isinstance(value, (int, float)):
        result = float(value)
        logger.debug(f"Monto convertido desde número: {result}")
        return result
    
    if isinstance(value, str):
        # Limpiar el string
        clean_value = value.strip().replace(',', '').replace('$', '').replace('€', '')
        
        # Manejar paréntesis como negativos (formato contable)
        if clean_value.startswith('(') and clean_value.endswith(')'):
            clean_value = '-' + clean_value[1:-1]
            logger.debug(f"Formato contable detectado, valor limpio: {clean_value}")
        
        try:
            result = float(clean_value)
            logger.debug(f"Monto parseado desde string: {result}")
            return result
        except ValueError as e:
            logger.error(f"Error parseando monto '{value}': {str(e)}")
            raise ValueError(f"Monto no válido: {value}")
    
    logger.error(f"Tipo de monto no soportado: {type(value)}")
    raise ValueError(f"Tipo de monto no soportado: {type(value)}")

def detect_excel_columns(worksheet):
    """Detecta automáticamente las columnas en la primera fila"""
    header_row = 1
    headers = {}
    
    # Mapeo de posibles nombres de columnas
    column_mappings = {
        'date': ['date', 'fecha', 'día', 'dia', 'transaction_date', 'fecha_transaccion'],
        'amount': ['amount', 'monto', 'valor', 'importe', 'cantidad', 'sum', 'total'],
        'description': ['description', 'descripcion', 'concepto', 'detalle', 'memo', 'detail'],
        'notes': ['notes', 'notas', 'observaciones', 'comentarios', 'remarks']
    }
    
    # Leer la primera fila para encontrar headers
    for col_idx, cell in enumerate(worksheet[header_row], 1):
        if cell.value:
            cell_value = str(cell.value).lower().strip()
            
            # Buscar matches en los mapeos
            for field, possible_names in column_mappings.items():
                for possible_name in possible_names:
                    if possible_name in cell_value:
                        headers[field] = col_idx
                        break
                if field in headers:
                    break
    
    return headers, header_row

def import_transactions_from_excel(
    db: Session, 
    user_id: int, 
    account_id: int, 
    file_content: bytes, 
    filename: str,
    allow_duplicates: bool = False  # NUEVO PARÁMETRO
) -> Dict[str, Any]:
    """Importa transacciones desde contenido Excel con prevención de duplicados"""
    
    logger.info(f"Iniciando importación de Excel para usuario {user_id}, cuenta {account_id}, archivo: {filename}")
    
    # Verificar que la cuenta pertenece al usuario
    account = db.query(Account).filter(
        Account.id == account_id,
        Account.user_id == user_id,
        Account.active == True
    ).first()
    
    if not account:
        logger.error(f"Cuenta {account_id} no encontrada o no autorizada para usuario {user_id}")
        raise ValueError("Cuenta no encontrada o no autorizada")
    
    logger.debug(f"Cuenta encontrada: {account.name} (ID: {account.id})")
    
    results = {
        'total_records': 0,
        'successful_imports': 0,
        'failed_imports': 0,
        'skipped_duplicates': 0,  # NUEVO CONTADOR
        'errors': []
    }
    
    try:
        # Cargar el workbook desde bytes
        logger.debug("Cargando workbook de Excel")
        workbook = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
        
        # Usar la primera hoja
        worksheet = workbook.active
        logger.debug(f"Hoja activa seleccionada: {worksheet.title if worksheet else 'None'}")
        
        # Detectar columnas automáticamente
        logger.debug("Detectando columnas automáticamente")
        column_map, header_row = detect_excel_columns(worksheet)
        logger.debug(f"Columnas detectadas: {column_map}")
        logger.debug(f"Fila de encabezados: {header_row}")
        
        if not column_map.get('date') or not column_map.get('amount'):
            error_msg = "No se pudieron detectar las columnas requeridas (fecha y monto). Verifica que el archivo tenga headers en la primera fila."
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        # Procesar filas
        logger.debug("Iniciando procesamiento de filas")
        if worksheet is None:
            logger.error("No se pudo obtener una hoja de trabajo válida del archivo Excel")
            raise ValueError("No se pudo obtener una hoja de trabajo válida del archivo Excel")
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=header_row + 1), start=header_row + 1):
            # Saltar filas vacías
            if all(cell.value is None or str(cell.value).strip() == '' for cell in row):
                logger.debug(f"Saltando fila vacía {row_idx}")
                continue
                
            results['total_records'] += 1
            logger.debug(f"Procesando fila {row_idx}")
            
            try:
                # Extraer valores de las celdas
                date_value = row[column_map['date'] - 1].value if column_map.get('date') else None
                amount_value = row[column_map['amount'] - 1].value if column_map.get('amount') else None
                description_value = row[column_map['description'] - 1].value if column_map.get('description') else ""
                notes_value = row[column_map['notes'] - 1].value if column_map.get('notes') else ""
                
                logger.debug(f"Fila {row_idx} - Valores extraídos: date={date_value}, amount={amount_value}, description='{description_value}'")
                
                # Validar y convertir valores
                transaction_date = parse_excel_date(date_value)
                amount = parse_excel_amount(amount_value)
                
                if transaction_date is None:
                    raise ValueError("Fecha requerida")
                
                if amount == 0:
                    raise ValueError("El monto no puede ser cero")
                
                # Redondear el monto a 2 decimales para evitar problemas de precisión
                amount = round(amount, 2)
                
                # Obtener un status_id válido
                default_status_id = get_default_transaction_status_id(db)
                # Asegurarse de que sea un int y no un SQLAlchemy Column
                if hasattr(default_status_id, 'id'):
                    default_status_id = default_status_id.id
                elif hasattr(default_status_id, 'value'):
                    default_status_id = default_status_id.value
                elif isinstance(default_status_id, int):
                    pass  # already int
                else:
                    # If it's a SQLAlchemy Column, raise an error or set a default
                    logger.error(f"default_status_id is not an int: {default_status_id} (type: {type(default_status_id)})")
                    raise ValueError("No se pudo obtener un status_id válido para la transacción")

                # Crear transacción
                transaction_data = TransactionCreateRequest(
                    amount=amount,
                    description=str(description_value) if description_value else "",
                    transaction_date=transaction_date,
                    account_id=account_id,
                    notes=str(notes_value) if notes_value else "",
                    status_id=default_status_id,
                    external_id=f"{filename}_{row_idx}"  # GENERAR external_id ÚNICO
                )
                
                # CREAR CON VALIDACIÓN DE DUPLICADOS
                try:
                    create_transaction(
                        db=db, 
                        user_id=user_id, 
                        transaction_data=transaction_data,
                        import_source=filename,
                        skip_duplicate_check=allow_duplicates
                    )
                    results['successful_imports'] += 1
                    logger.debug(f"Fila {row_idx} importada exitosamente")
                    
                except ValueError as e:
                    if "duplicada" in str(e):
                        results['skipped_duplicates'] += 1
                        logger.info(f"Fila {row_idx} saltada (duplicado): {str(e)}")
                        if not allow_duplicates:
                            continue  # Saltar sin contar como error
                    raise  # Re-lanzar otros errores
                
            except Exception as e:
                results['failed_imports'] += 1
                error_msg = f"Fila {row_idx}: {str(e)}"
                results['errors'].append(error_msg)
                logger.warning(error_msg)
                continue
        
        workbook.close()
        logger.info(f"Importación completada: {results['successful_imports']} exitosas, {results['skipped_duplicates']} duplicados saltados, {results['failed_imports']} fallidas de {results['total_records']} total")
                
    except Exception as e:
        error_msg = f"Error procesando archivo Excel: {str(e)}"
        logger.error(error_msg)
        raise ValueError(error_msg)
    
    return results

# Mantener función legacy para CSV si es necesario
def import_transactions_from_csv(
    db: Session, 
    user_id: int, 
    account_id: int, 
    csv_content: str, 
    filename: str
) -> Dict[str, Any]:
    """Función legacy para CSV - redirige a Excel si el archivo es Excel"""
    if filename.lower().endswith(('.xlsx', '.xls')):
        raise ValueError("Use import_transactions_from_excel para archivos Excel")
    
    # Implementación CSV original...
    return import_transactions_from_excel(db, user_id, account_id, csv_content.encode(), filename)

def import_transactions_with_profile(
    db: Session,
    user_id: int,
    profile_id: int,
    file_content: bytes,
    filename: str
) -> Dict[str, Any]:
    """Importa transacciones usando un perfil de importación configurado"""
    
    logger.info(f"Iniciando importación con perfil para usuario {user_id}, perfil {profile_id}, archivo: {filename}")
    logger.debug(f"Tamaño del archivo: {len(file_content)} bytes")
    
    # Obtener el perfil de importación
    profile = db.query(FileImportProfile).filter(
        FileImportProfile.id == profile_id,
        FileImportProfile.user_id == user_id
    ).first()
    
    if not profile:
        logger.error(f"Perfil de importación {profile_id} no encontrado para usuario {user_id}")
        raise ValueError("Perfil de importación no encontrado")
    
    logger.debug(f"Perfil encontrado: {profile.name} (ID: {profile.id}), cuenta: {profile.account_id}")
    
    # Verificar que la cuenta pertenece al usuario
    account = db.query(Account).filter(
        Account.id == profile.account_id,
        Account.user_id == user_id,
        Account.active == True
    ).first()
    
    if not account:
        logger.error(f"Cuenta {profile.account_id} no encontrada o no autorizada para usuario {user_id}")
        raise ValueError("Cuenta no encontrada o no autorizada")
    
    logger.debug(f"Cuenta encontrada: {account.name} (ID: {account.id})")
    
    # Obtener los mapeos de columnas
    column_mappings = db.query(FileColumnMapping).filter(
        FileColumnMapping.profile_id == profile_id
    ).order_by(FileColumnMapping.position).all()
    
    if not column_mappings:
        logger.error(f"No se encontraron mapeos de columnas para el perfil {profile_id}")
        raise ValueError("No se encontraron mapeos de columnas para el perfil")
    
    logger.debug(f"Mapeos de columnas encontrados: {len(column_mappings)} mapeos")
    for mapping in column_mappings:
        logger.debug(f"  - {mapping.target_field_name}: columna {mapping.source_column_index} / '{mapping.source_column_name}'")
    
    results = {
        'total_records': 0,
        'successful_imports': 0,
        'failed_imports': 0,
        'errors': []
    }
    
    try:
        # Verificar que el archivo no esté vacío
        if not file_content:
            logger.error("El archivo está vacío")
            raise ValueError("El archivo está vacío")
        
        logger.debug(f"Procesando archivo: {filename}, tamaño: {len(file_content)} bytes")
        
        # Determinar el tipo de archivo y procesarlo
        filename_lower = filename.lower()
        if filename_lower.endswith('.csv'):
            logger.info("Procesando como CSV")
            return _process_csv_with_profile(db, user_id, profile, column_mappings, file_content, results)
        elif filename_lower.endswith(('.xlsx', '.xls')):
            logger.info("Procesando como Excel")
            return _process_excel_with_profile(db, user_id, profile, column_mappings, file_content, results)
        else:
            # Intentar detectar por contenido si la extensión no es clara
            logger.debug("Extensión no reconocida, intentando detectar por contenido")
            try:
                # Intentar como texto (CSV)
                content_preview = file_content[:1024].decode('utf-8', errors='ignore')
                if any(delimiter in content_preview for delimiter in [',', ';', '\t']):
                    logger.info("Detectado como CSV por contenido")
                    return _process_csv_with_profile(db, user_id, profile, column_mappings, file_content, results)
            except Exception as e:
                logger.debug(f"Error detectando como CSV: {str(e)}")
                pass
            
            # Si no se puede detectar, asumir Excel
            logger.info("Procesando como Excel por defecto")
            return _process_excel_with_profile(db, user_id, profile, column_mappings, file_content, results)
            
    except Exception as e:
        error_msg = f"Error procesando archivo: {str(e)}"
        logger.error(f"Error en import_transactions_with_profile: {str(e)}")
        raise ValueError(error_msg)

def _process_csv_with_profile(
    db: Session,
    user_id: int,
    profile: FileImportProfile,
    column_mappings: List[FileColumnMapping],
    file_content: bytes,
    results: Dict[str, Any]
) -> Dict[str, Any]:
    """Procesa un archivo CSV usando el perfil de importación"""
    
    logger.info(f"Procesando CSV para usuario {user_id} con perfil {profile.id}")
    
    try:
        # Decodificar el contenido
        encoding = getattr(profile, 'encoding', 'utf-8') or 'utf-8'
        logger.debug(f"Decodificando CSV con encoding: {encoding}")
        csv_content = file_content.decode(encoding)
        
        # Crear un reader CSV
        delimiter = getattr(profile, 'delimiter', ',') or ','
        logger.debug(f"Usando delimitador: '{delimiter}'")
        csv_reader = csv.reader(
            io.StringIO(csv_content),
            delimiter=delimiter
        )
        
        rows = list(csv_reader)
        logger.debug(f"CSV cargado: {len(rows)} filas totales")
        
        # Saltar headers si existen
        has_header = getattr(profile, 'has_header', True)
        start_row = 1 if has_header else 0
        logger.debug(f"Tiene headers: {has_header}, fila de inicio: {start_row}")
        
        # Crear mapeo de columnas por nombre o índice
        column_map = {}
        for mapping in column_mappings:
            source_column_name = getattr(mapping, 'source_column_name', None)
            source_column_index = getattr(mapping, 'source_column_index', None)
            target_field_name = getattr(mapping, 'target_field_name', '')
            
            if has_header and source_column_name:
                # Buscar el índice de la columna por nombre
                try:
                    header_row = rows[0]
                    column_index = header_row.index(source_column_name)
                    column_map[target_field_name] = column_index
                    logger.debug(f"Mapeo por nombre: {target_field_name} -> columna {column_index} ('{source_column_name}')")
                except (IndexError, ValueError):
                    # Si no se encuentra, usar el índice configurado
                    if source_column_index is not None:
                        column_map[target_field_name] = source_column_index
                        logger.debug(f"Mapeo por índice (fallback): {target_field_name} -> columna {source_column_index}")
            else:
                # Usar el índice configurado
                if source_column_index is not None:
                    column_map[target_field_name] = source_column_index
                    logger.debug(f"Mapeo por índice: {target_field_name} -> columna {source_column_index}")
        
        logger.debug(f"Mapeo final de columnas: {column_map}")
        
        # Procesar filas de datos
        data_rows = rows[start_row:]
        logger.debug(f"Procesando {len(data_rows)} filas de datos")
        
        for row_idx, row in enumerate(data_rows, start=start_row + 1):
            if not row or all(not cell.strip() for cell in row):
                logger.debug(f"Saltando fila vacía {row_idx}")
                continue
                
            results['total_records'] += 1
            logger.debug(f"Procesando fila {row_idx}: {row}")
            
            try:
                transaction_data = _extract_transaction_data(db, row, column_map, profile, row_idx)
                account_id = getattr(profile, 'account_id')
                transaction_data.account_id = account_id
                
                create_transaction(db, user_id, transaction_data)
                results['successful_imports'] += 1
                logger.debug(f"Fila {row_idx} importada exitosamente")
                
            except Exception as e:
                results['failed_imports'] += 1
                error_msg = f"Fila {row_idx}: {str(e)}"
                results['errors'].append(error_msg)
                logger.warning(f"Error en fila {row_idx}: {str(e)}")
                continue
                
        logger.info(f"CSV procesado: {results['successful_imports']} exitosas, {results['failed_imports']} fallidas de {results['total_records']} total")
                
    except Exception as e:
        error_msg = f"Error procesando CSV: {str(e)}"
        logger.error(error_msg)
        raise ValueError(error_msg)
    
    return results

def _process_excel_with_profile(
    db: Session,
    user_id: int,
    profile: FileImportProfile,
    column_mappings: List[FileColumnMapping],
    file_content: bytes,
    results: Dict[str, Any]
) -> Dict[str, Any]:
    """Procesa un archivo Excel usando el perfil de importación - soporta .xlsx y .xls"""
    
    logger.info(f"Procesando Excel para usuario {user_id} con perfil {profile.id}")
    
    try:
        # Verificar que el contenido no esté vacío
        if not file_content:
            logger.error("El archivo está vacío")
            raise ValueError("El archivo está vacío")
        
        # Verificar tamaño mínimo para un archivo Excel válido
        if len(file_content) < 100:
            logger.error(f"Archivo demasiado pequeño: {len(file_content)} bytes")
            raise ValueError("El archivo es demasiado pequeño para ser un Excel válido")
        
        # Intentar cargar primero como .xlsx, luego como .xls
        worksheet = None
        workbook = None
        is_xls_format = False
        
        try:
            logger.debug(f"Intentando cargar como Excel .xlsx, primeros 50 bytes: {file_content[:50]}")
            workbook = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
            logger.info(f"Excel .xlsx cargado exitosamente, hojas disponibles: {workbook.sheetnames}")
            
            # Seleccionar la hoja
            sheet_name = getattr(profile, 'sheet_name', None)
            if sheet_name:
                try:
                    worksheet = workbook[sheet_name]
                    logger.debug(f"Hoja seleccionada: {sheet_name}")
                except KeyError:
                    available_sheets = workbook.sheetnames
                    error_msg = f"Hoja '{sheet_name}' no encontrada. Hojas disponibles: {', '.join(available_sheets)}"
                    logger.error(error_msg)
                    raise ValueError(error_msg)
            else:
                worksheet = workbook.active
                logger.debug(f"Hoja activa seleccionada: {worksheet.title if worksheet else 'None'}")
                
        except Exception as xlsx_error:
            logger.warning(f"Error cargando como .xlsx: {str(xlsx_error)}")
            try:
                logger.debug("Intentando cargar como Excel .xls...")
                workbook = xlrd.open_workbook(file_contents=file_content)
                is_xls_format = True
                
                # Para .xls, seleccionar la hoja
                sheet_name = getattr(profile, 'sheet_name', None)
                if sheet_name:
                    try:
                        worksheet = workbook.sheet_by_name(sheet_name)
                        logger.debug(f"Hoja .xls seleccionada: {sheet_name}")
                    except Exception:  # xlrd puede lanzar diferentes tipos de errores
                        available_sheets = workbook.sheet_names()
                        error_msg = f"Hoja '{sheet_name}' no encontrada. Hojas disponibles: {', '.join(available_sheets)}"
                        logger.error(error_msg)
                        raise ValueError(error_msg)
                else:
                    worksheet = workbook.sheet_by_index(0)
                    logger.debug("Hoja activa .xls seleccionada (índice 0)")
                    
                logger.info("Excel .xls cargado exitosamente")
                
            except Exception as xls_error:
                logger.error(f"Error cargando como .xls: {str(xls_error)}")
                # Si falla como Excel, intentar detectar si es realmente un CSV
                try:
                    content_str = file_content.decode('utf-8')
                    if ',' in content_str or ';' in content_str:
                        error_msg = "El archivo parece ser CSV, no Excel. Use el endpoint de CSV o cambie la extensión del archivo."
                        logger.error(error_msg)
                        raise ValueError(error_msg)
                except UnicodeDecodeError:
                    pass
                error_msg = f"Archivo Excel inválido. Errores: .xlsx: {str(xlsx_error)}, .xls: {str(xls_error)}"
                logger.error(error_msg)
                raise ValueError(error_msg)
            
        if worksheet is None:
            logger.error("No se pudo obtener una hoja de trabajo válida")
            raise ValueError("No se pudo obtener una hoja de trabajo válida del archivo Excel")
        
        # Crear mapeo de columnas
        logger.debug("Creando mapeo de columnas")
        column_map = {}
        has_header = getattr(profile, 'has_header', True)
        header_row_num = getattr(profile, 'header_row', 1)
        logger.debug(f"Configuración: has_header={has_header}, header_row={header_row_num}")
        
        if has_header and header_row_num:
            # Leer la fila de headers según el formato
            header_row = None
            
            if is_xls_format:
                # Para .xls usar xlrd (índices 0-based)
                header_row_idx = header_row_num - 1
                logger.debug(f"Leyendo headers .xls en fila {header_row_idx}")
                # Usar hasattr para verificar si existen los atributos antes de usarlos
                if hasattr(worksheet, 'nrows') and hasattr(worksheet, 'ncols') and hasattr(worksheet, 'cell_value'):
                    if header_row_idx < worksheet.nrows:  # type: ignore
                        header_row = [worksheet.cell_value(header_row_idx, col) for col in range(worksheet.ncols)]  # type: ignore
                        logger.debug(f"Headers .xls encontrados: {header_row}")
            else:
                # Para .xlsx usar openpyxl
                logger.debug(f"Leyendo headers .xlsx en fila {header_row_num}")
                if hasattr(worksheet, 'iter_rows'):
                    header_row = list(worksheet.iter_rows(  # type: ignore
                        min_row=header_row_num,
                        max_row=header_row_num,
                        values_only=True
                    ))[0]
                    logger.debug(f"Headers .xlsx encontrados: {header_row}")
            
            if header_row:
                for mapping in column_mappings:
                    source_column_name = getattr(mapping, 'source_column_name', None)
                    source_column_index = getattr(mapping, 'source_column_index', None)
                    target_field_name = getattr(mapping, 'target_field_name', '')
                    
                    if source_column_name:
                        try:
                            column_index = header_row.index(source_column_name)
                            column_map[target_field_name] = column_index
                            logger.debug(f"Mapeo por nombre: {target_field_name} -> columna {column_index} ('{source_column_name}')")
                        except ValueError:
                            if source_column_index is not None:
                                column_map[target_field_name] = source_column_index
                                logger.debug(f"Mapeo por índice (fallback): {target_field_name} -> columna {source_column_index}")
                    else:
                        if source_column_index is not None:
                            column_map[target_field_name] = source_column_index
                            logger.debug(f"Mapeo por índice: {target_field_name} -> columna {source_column_index}")
        else:
            # Usar índices configurados
            logger.debug("Usando mapeo por índices (sin headers)")
            for mapping in column_mappings:
                source_column_index = getattr(mapping, 'source_column_index', None)
                target_field_name = getattr(mapping, 'target_field_name', '')
                if source_column_index is not None:
                    column_map[target_field_name] = source_column_index
                    logger.debug(f"Mapeo por índice: {target_field_name} -> columna {source_column_index}")
        
        logger.debug(f"Mapeo final de columnas: {column_map}")
        
        # Procesar filas de datos
        start_row_num = getattr(profile, 'start_row', None)
        if start_row_num is None:
            start_row_num = header_row_num + 1 if header_row_num else 1
        
        skip_empty_rows = getattr(profile, 'skip_empty_rows', True)
        logger.debug(f"Configuración procesamiento: start_row={start_row_num}, skip_empty_rows={skip_empty_rows}")
        
        if is_xls_format:
            # Procesar filas para .xls
            start_row_idx = start_row_num - 1  # xlrd usa índices 0-based
            logger.debug(f"Procesando filas .xls desde índice {start_row_idx}")
            
            if hasattr(worksheet, 'nrows') and hasattr(worksheet, 'ncols') and hasattr(worksheet, 'cell_value'):
                total_rows = worksheet.nrows  # type: ignore
                logger.debug(f"Total de filas en .xls: {total_rows}")
                
                for row_idx in range(start_row_idx, total_rows):
                    row = [worksheet.cell_value(row_idx, col) for col in range(worksheet.ncols)]  # type: ignore
                    
                    if skip_empty_rows and (not row or all(cell is None or str(cell).strip() == '' for cell in row)):
                        logger.debug(f"Saltando fila vacía {row_idx + 1}")
                        continue
                        
                    results['total_records'] += 1
                    logger.debug(f"Procesando fila .xls {row_idx + 1}: {row}")
                    
                    try:
                        transaction_data = _extract_transaction_data(db, row, column_map, profile, row_idx + 1)
                        account_id = getattr(profile, 'account_id')
                        transaction_data.account_id = account_id
                        
                        create_transaction(db, user_id, transaction_data)
                        results['successful_imports'] += 1
                        logger.debug(f"Fila .xls {row_idx + 1} importada exitosamente")
                        
                    except Exception as e:
                        results['failed_imports'] += 1
                        error_msg = f"Fila {row_idx + 1}: {str(e)}"
                        results['errors'].append(error_msg)
                        logger.warning(f"Error en fila .xls {row_idx + 1}: {str(e)}")
                        continue
        else:
            # Procesar filas para .xlsx
            logger.debug(f"Procesando filas .xlsx desde fila {start_row_num}")
            if hasattr(worksheet, 'iter_rows'):
                for row_idx, row in enumerate(worksheet.iter_rows(  # type: ignore
                    min_row=start_row_num,
                    values_only=True
                ), start=start_row_num):
                    
                    if skip_empty_rows and (not row or all(cell is None or str(cell).strip() == '' for cell in row)):
                        logger.debug(f"Saltando fila vacía {row_idx}")
                        continue
                        
                    results['total_records'] += 1
                    logger.debug(f"Procesando fila .xlsx {row_idx}: {row}")
                    
                    try:
                        transaction_data = _extract_transaction_data(db, list(row), column_map, profile, row_idx)
                        account_id = getattr(profile, 'account_id')
                        transaction_data.account_id = account_id
                        
                        create_transaction(db, user_id, transaction_data)
                        results['successful_imports'] += 1
                        logger.debug(f"Fila .xlsx {row_idx} importada exitosamente")
                        
                    except Exception as e:
                        results['failed_imports'] += 1
                        error_msg = f"Fila {row_idx}: {str(e)}"
                        results['errors'].append(error_msg)
                        logger.warning(f"Error en fila .xlsx {row_idx}: {str(e)}")
                        continue
        
        # Cerrar el workbook si es openpyxl
        if not is_xls_format and hasattr(workbook, 'close'):
            workbook.close()  # type: ignore
            logger.debug("Workbook .xlsx cerrado")
        
        logger.info(f"Excel procesado: {results['successful_imports']} exitosas, {results['failed_imports']} fallidas de {results['total_records']} total")
        
    except Exception as e:
        error_msg = f"Error procesando Excel: {str(e)}"
        logger.error(error_msg)
        raise ValueError(error_msg)
    
    return results

def _extract_transaction_data(
    db: Session,
    row: List,
    column_map: Dict[str, int],
    profile: FileImportProfile,
    row_idx: int
) -> TransactionCreateRequest:
    """Extrae los datos de transacción de una fila usando el mapeo de columnas"""
    
    logger.debug(f"Extrayendo datos de transacción de fila {row_idx}")
    
    def get_cell_value(field_name: str, default=None):
        if field_name in column_map and column_map[field_name] < len(row):
            value = row[column_map[field_name]]
            logger.debug(f"  {field_name}: {value} (columna {column_map[field_name]})")
            return value if value is not None else default
        logger.debug(f"  {field_name}: {default} (no mapeado o fuera de rango)")
        return default
    
    # Extraer fecha
    date_value = get_cell_value('date')
    if not date_value:
        logger.error(f"Fila {row_idx}: Fecha requerida pero no encontrada")
        raise ValueError("Fecha requerida")
    
    transaction_date = parse_excel_date(date_value)
    if transaction_date is None:
        logger.error(f"Fila {row_idx}: Fecha inválida: {date_value}")
        raise ValueError("Fecha inválida")
    
    logger.debug(f"  Fecha procesada: {transaction_date}")
    
    # Extraer montos - detectar automáticamente el esquema disponible
    amount = 0.0
    amount_schema = getattr(profile, 'amount_schema')
    
    if hasattr(amount_schema, 'value'):
        schema_value = amount_schema.value
    else:
        schema_value = str(amount_schema)
    
    logger.debug(f"  Esquema configurado: {schema_value}")
    
    # Detectar qué tipos de columnas están disponibles
    has_single_amount = 'amount' in column_map
    has_separate_amounts = ('expense_amount' in column_map and 'income_amount' in column_map) or \
                          ('debit_amount' in column_map and 'credit_amount' in column_map)
    
    logger.debug(f"  Columnas disponibles: single_amount={has_single_amount}, separate_amounts={has_separate_amounts}")
    logger.debug(f"  Mapeo de columnas: {list(column_map.keys())}")
    
    if has_separate_amounts:
        # Usar columnas separadas (prioritario)
        logger.debug("  Usando lógica de columnas separadas")
        
        # Primero intentar con expense_amount/income_amount
        expense_amount = parse_excel_amount(get_cell_value('expense_amount', 0))
        income_amount = parse_excel_amount(get_cell_value('income_amount', 0))
        
        # Si no están disponibles, usar debit_amount/credit_amount
        if expense_amount == 0 and income_amount == 0:
            debit_amount = parse_excel_amount(get_cell_value('debit_amount', 0))
            credit_amount = parse_excel_amount(get_cell_value('credit_amount', 0))
            
            debit_column_is_expense = getattr(profile, 'debit_column_is_expense', True)
            
            logger.debug(f"  Montos débito/crédito: debit={debit_amount}, credit={credit_amount}")
            logger.debug(f"  debit_column_is_expense: {debit_column_is_expense}")
            
            if debit_amount != 0:
                amount = -abs(debit_amount) if debit_column_is_expense else abs(debit_amount)
                logger.debug(f"  Usando debit_amount: {amount}")
            elif credit_amount != 0:
                amount = abs(credit_amount) if not debit_column_is_expense else -abs(credit_amount)
                logger.debug(f"  Usando credit_amount: {amount}")
        else:
            logger.debug(f"  Montos ingreso/gasto: expense={expense_amount}, income={income_amount}")
            
            if expense_amount != 0:
                amount = -abs(expense_amount)  # Gastos siempre negativos
                logger.debug(f"  Usando expense_amount: {amount}")
            elif income_amount != 0:
                amount = abs(income_amount)  # Ingresos siempre positivos
                logger.debug(f"  Usando income_amount: {amount}")
                
    elif has_single_amount:
        # Usar columna única
        logger.debug("  Usando lógica de columna única")
        amount_value = get_cell_value('amount', 0)
        amount = parse_excel_amount(amount_value)
        
        logger.debug(f"  Monto único obtenido: {amount}")
        
        # Aplicar reglas de interpretación
        positive_is_income = getattr(profile, 'positive_is_income', True)
        if not positive_is_income:
            amount = -amount  # Invertir signo si positivo no es ingreso
            logger.debug(f"  Signo invertido por configuración: {amount}")
    
    else:
        # No se encontraron columnas de monto válidas
        logger.error(f"  No se encontraron columnas de monto válidas en el mapeo")
        logger.error(f"  Columnas mapeadas: {list(column_map.keys())}")
        raise ValueError("No se encontraron columnas de monto válidas en la configuración")
    
    if amount == 0:
        logger.error(f"Fila {row_idx}: El monto no puede ser cero")
        raise ValueError("El monto no puede ser cero")
    
    # Extraer otros campos
    description = str(get_cell_value('description', '')).strip()
    notes = str(get_cell_value('notes', '')).strip()
    
    logger.debug(f"  Monto final: {amount}")
    logger.debug(f"  Descripción: '{description}'")
    logger.debug(f"  Notas: '{notes}'")
    
    # Redondear el monto a 2 decimales para evitar problemas de precisión
    amount = round(amount, 2)
    
    # Obtener un status_id válido  
    default_status_id = get_default_transaction_status_id(db)
    # Asegurarse de que sea un int
    if hasattr(default_status_id, 'id'):
        default_status_id = default_status_id.id
    elif hasattr(default_status_id, 'value'):
        default_status_id = default_status_id.value
    elif isinstance(default_status_id, int):
        pass  # already int
    else:
        raise ValueError("No se pudo obtener un status_id válido para la transacción")

    transaction_request = TransactionCreateRequest(
        amount=amount,
        description=description,
        transaction_date=transaction_date,
        account_id=0,  # Se asignará después
        notes=notes,
        status_id=default_status_id,  # Usar status válido
        # No incluir campos opcionales que no vienen del archivo
        subcategory_id=None,
        envelope_id=None,
        transfer_account_id=None,
        is_recurring=False,
        is_planned=False,
        kakebo_emotion=None,
        external_id=None
    )
    
    logger.debug(f"Transacción extraída exitosamente de fila {row_idx}")
    return transaction_request